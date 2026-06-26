"""Prueba: el nombre del profesional se recupera de otra hoja si falta en la del mes."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from ods_reporter.infrastructure.excel.openpyxl_reader import OpenpyxlExcelReader


def _build_sheet(ws, *, professional: str) -> None:
    """Crea una hoja con la cabecera mínima y la tabla de actividades."""
    ws["I4"] = "PROFESIONAL RESPONSABLE"
    if professional:
        ws["K4"] = professional
    # Encabezado de la tabla (fila 9) con las columnas requeridas.
    ws["A9"] = "ID"
    ws["B9"] = "ACTIVIDADES"
    ws["C9"] = "ENTREGABLES"
    ws["F9"] = "DESCRIPCIÓN DE ACTIVIDADES O ENTREGABLES REALIZADOS EN EL PERIODO"
    # Una actividad de ejemplo.
    ws["A10"] = 1
    ws["B10"] = "i. Actividad de prueba"
    ws["C10"] = "Entregable de prueba"
    ws["F10"] = "No se requirió esta actividad para el periodo reportado."


def test_name_recovered_from_other_sheet(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    enero = wb.active
    enero.title = "ENERO"
    _build_sheet(enero, professional="Paola Andrea Moya")  # nombre solo en enero
    mayo = wb.create_sheet("MAYO")
    _build_sheet(mayo, professional="")  # MAYO sin nombre

    path = tmp_path / "sin_nombre_mayo.xlsx"
    wb.save(str(path))

    report = OpenpyxlExcelReader().read_month(path, "MAYO")
    assert report.professional_name == "Paola Andrea Moya"


def test_name_empty_when_no_sheet_has_it(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    mayo = wb.active
    mayo.title = "MAYO"
    _build_sheet(mayo, professional="")
    path = tmp_path / "sin_nombre.xlsx"
    wb.save(str(path))

    report = OpenpyxlExcelReader().read_month(path, "MAYO")
    assert report.professional_name == ""
