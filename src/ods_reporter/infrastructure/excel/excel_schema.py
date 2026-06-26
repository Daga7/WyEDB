"""Esquema y detección de la estructura del Excel de reportes ODS.

Centraliza el conocimiento de "dónde está cada cosa" de forma **robusta**: en
lugar de coordenadas fijas (frágiles si el formato cambia), se detecta la fila de
encabezados y se mapean las columnas por su texto. Así, si las columnas se
desplazan, el lector sigue funcionando.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import TYPE_CHECKING

from ods_reporter.domain.entities.ods_metadata import ODSMetadata
from ods_reporter.domain.exceptions import ExcelReadError
from ods_reporter.shared.text_utils import normalize_text

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# --- Textos de encabezado (normalizados) usados para localizar columnas ---
_HEADER_ID = "id"
_HEADER_ACTIVITIES = "actividades"
_HEADER_ENTREGABLES = "entregables"
_HEADER_DESCRIPTION = "descripcion de actividades o entregables realizados"

# --- Marcadores que indican el FIN de la tabla de actividades ---
# Independientes de la empresa (sirven para Ecopetrol, HOCOL, etc.).
_STOP_MARKERS = (
    "otras actividades solicitadas",
    "relacion de comisiones",
    "fin del reporte",
)

# Hasta qué fila buscar el encabezado y los metadatos de cabecera.
_MAX_HEADER_SCAN_ROW = 30

# Etiquetas de metadatos (normalizadas) -> campo de ODSMetadata.
_METADATA_LABELS: dict[str, tuple[str, ...]] = {
    "ods_number": ("ods n",),
    "responsible_professional": ("profesional responsable",),
    "client": ("nombre del cliente",),
    "contract_number": ("n° contrato", "no contrato", "n contrato"),
    "reported_period": ("periodo reportado",),
}


@dataclass(frozen=True, slots=True)
class TableLayout:
    """Posiciones detectadas de la tabla de actividades (índices 1-based).

    ``entregables_col`` es opcional: algunas plantillas (p. ej. HOCOL/ODS 266) no
    tienen columna ENTREGABLES; en ese caso cada actividad es su propio entregable.
    """

    header_row: int
    id_col: int
    activities_col: int
    description_col: int
    entregables_col: int | None = None


def find_table_layout(ws: Worksheet) -> TableLayout:
    """Detecta la fila de encabezados y las columnas clave de la tabla.

    Raises
    ------
    ExcelReadError
        Si no se encuentra una fila de encabezados reconocible o faltan columnas
        imprescindibles (ID, ACTIVIDADES y la de descripción).
    """
    header_row = _find_header_row(ws)
    columns = _map_header_columns(ws, header_row)

    id_col = columns.get(_HEADER_ID)
    activities_col = columns.get(_HEADER_ACTIVITIES)
    entregables_col = columns.get(_HEADER_ENTREGABLES)  # opcional
    description_col = _find_column_by_prefix(columns, _HEADER_DESCRIPTION)

    missing = [
        name
        for name, value in (
            ("ID", id_col),
            ("ACTIVIDADES", activities_col),
            ("DESCRIPCIÓN DE ACTIVIDADES...", description_col),
        )
        if value is None
    ]
    if missing:
        raise ExcelReadError(
            "No se reconocieron las columnas requeridas en el encabezado: "
            + ", ".join(missing)
        )

    assert id_col is not None and activities_col is not None and description_col is not None
    return TableLayout(
        header_row=header_row,
        id_col=id_col,
        activities_col=activities_col,
        description_col=description_col,
        entregables_col=entregables_col,
    )


def is_stop_marker(text: str) -> bool:
    """``True`` si el texto (ya normalizado o no) indica el fin de la tabla."""
    normalized = normalize_text(text)
    return any(marker in normalized for marker in _STOP_MARKERS)


def find_professional_name(ws: Worksheet) -> str:
    """Lee el nombre del profesional responsable de una hoja (o '' si no está)."""
    return _find_value_right_of_label(
        ws, _METADATA_LABELS["responsible_professional"], _MAX_HEADER_SCAN_ROW
    )


def read_metadata(ws: Worksheet, header_row: int, source_file: str) -> ODSMetadata:
    """Extrae los metadatos de cabecera (best-effort; campos faltantes -> '')."""
    values: dict[str, str] = {}
    for field_name, keys in _METADATA_LABELS.items():
        values[field_name] = _find_value_right_of_label(ws, keys, header_row)

    return ODSMetadata(
        ods_number=values["ods_number"],
        contract_number=values["contract_number"],
        client=values["client"],
        responsible_professional=values["responsible_professional"],
        reported_period=values["reported_period"],
        source_file=source_file,
    )


# --- Funciones internas ---

def _find_header_row(ws: Worksheet) -> int:
    limit = min(_MAX_HEADER_SCAN_ROW, ws.max_row)
    for row in range(1, limit + 1):
        a = normalize_text(str(ws.cell(row=row, column=1).value or ""))
        b = normalize_text(str(ws.cell(row=row, column=2).value or ""))
        if a == _HEADER_ID and b == _HEADER_ACTIVITIES:
            return row
    raise ExcelReadError(
        "No se encontró la fila de encabezados de la tabla (ID / ACTIVIDADES)."
    )


def _map_header_columns(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        text = normalize_text(str(ws.cell(row=header_row, column=col).value or ""))
        if text and text not in mapping:
            mapping[text] = col
    return mapping


def _find_column_by_prefix(columns: dict[str, int], prefix: str) -> int | None:
    for text, col in columns.items():
        if text.startswith(prefix):
            return col
    return None


def _find_value_right_of_label(
    ws: Worksheet, label_keys: tuple[str, ...], max_row: int
) -> str:
    limit = min(max_row, ws.max_row)
    for row in range(1, limit + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_text(str(ws.cell(row=row, column=col).value or ""))
            if text and any(text.startswith(key) for key in label_keys):
                value = _first_non_empty_to_right(ws, row, col)
                if value:
                    return value
    return ""


def _first_non_empty_to_right(ws: Worksheet, row: int, start_col: int) -> str:
    for col in range(start_col + 1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is not None and str(value).strip():
            return _format_value(value)
    return ""


def _format_value(value: object) -> str:
    if isinstance(value, datetime | date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()
