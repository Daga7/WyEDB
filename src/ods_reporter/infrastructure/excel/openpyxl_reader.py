"""Implementación del lector de Excel con openpyxl.

Lee la hoja de un mes y extrae los metadatos de cabecera y las actividades
(agrupando las que abarcan varias sub-filas). Hace una extracción fiel: no
filtra ni interpreta el contenido (eso lo hace el normalizador en la Fase 5).
"""

from __future__ import annotations

import dataclasses
import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl

from ods_reporter.application.ports.excel_reader_port import (
    RawActivity,
    RawEntregable,
    RawOtherActivity,
    RawReport,
)
from ods_reporter.domain.exceptions import ExcelReadError, InvalidInputError
from ods_reporter.infrastructure.excel.excel_schema import (
    TableLayout,
    find_professional_name,
    find_table_layout,
    is_stop_marker,
    read_metadata,
)
from ods_reporter.shared.constants import EXCEL_EXTENSIONS, MONTHS
from ods_reporter.shared.text_utils import normalize_text

# Marcador de la sección de actividades adicionales ("OTRAS ACTIVIDADES
# SOLICITADAS POR ECOPETROL/HOCOL/..."): su fila es también el encabezado.
_OTHER_ACTIVITIES_MARKER = "otras actividades solicitadas"
# Encabezado (normalizado) de la columna de fecha dentro de esa sección.
_OTHER_DATE_HEADER = "fecha de ejecucion"
# Marcadores que terminan la sección de actividades adicionales.
_OTHER_END_MARKERS = ("fin del reporte", "relacion de comisiones")

logger = logging.getLogger(__name__)


class OpenpyxlExcelReader:
    """Lector de Excel basado en openpyxl (implementa ``ExcelReaderPort``)."""

    def read_month(self, file_path: Path, month: str) -> RawReport:
        self._validate_file(file_path)

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as exc:  # openpyxl lanza varios tipos según el problema
            raise ExcelReadError(
                f"No se pudo abrir el archivo Excel '{file_path.name}': {exc}"
            ) from exc

        try:
            sheet = self._select_sheet(workbook, month, file_path)
            layout = find_table_layout(sheet)
            metadata = read_metadata(sheet, layout.header_row, source_file=file_path.name)
            # Si el nombre del profesional falta en la hoja del mes, se busca en
            # las demás hojas (algunos profesionales solo lo diligencian en una).
            if not metadata.responsible_professional:
                fallback = self._find_name_in_other_sheets(workbook, month)
                if fallback:
                    metadata = dataclasses.replace(
                        metadata, responsible_professional=fallback
                    )
            activities, stop_row = self._read_activities(sheet, layout)
            other_activities = self._read_other_activities(sheet, layout, stop_row)
        finally:
            workbook.close()

        logger.info(
            "Excel '%s' [%s]: %d actividades y %d adicionales (profesional: %s)",
            file_path.name,
            month,
            len(activities),
            len(other_activities),
            metadata.responsible_professional or "desconocido",
        )
        return RawReport(
            metadata=metadata, activities=activities, other_activities=other_activities
        )

    # --- Validación ---

    @staticmethod
    def _validate_file(file_path: Path) -> None:
        if not file_path.exists():
            raise InvalidInputError(f"El archivo no existe: {file_path}")
        if file_path.suffix.lower() not in EXCEL_EXTENSIONS:
            raise InvalidInputError(
                f"Extensión no válida '{file_path.suffix}'. "
                f"Se esperaba una de: {', '.join(EXCEL_EXTENSIONS)}"
            )

    @staticmethod
    def _find_name_in_other_sheets(workbook: openpyxl.Workbook, current_month: str) -> str:
        """Busca el nombre del profesional en las demás hojas-mes del libro."""
        target = current_month.strip().lower()
        for name in workbook.sheetnames:
            if name.strip().lower() == target:
                continue
            if name.strip().upper() not in MONTHS:
                continue
            found = find_professional_name(workbook[name])
            if found:
                return found
        return ""

    @staticmethod
    def _select_sheet(
        workbook: openpyxl.Workbook, month: str, file_path: Path
    ) -> "openpyxl.worksheet.worksheet.Worksheet":
        # 1) Coincidencia exacta (tolerante a mayúsculas/espacios).
        target = month.strip().lower()
        for name in workbook.sheetnames:
            if name.strip().lower() == target:
                return workbook[name]
        # 2) Coincidencia parcial: el mes está contenido en el nombre de la hoja
        #    (p. ej. "MARZO" dentro de "FEB_MARZO"), o viceversa.
        for name in workbook.sheetnames:
            normalized = name.strip().lower()
            if target and (target in normalized or normalized in target):
                return workbook[name]
        raise ExcelReadError(
            f"El archivo '{file_path.name}' no contiene la hoja del mes '{month}'. "
            f"Hojas disponibles: {', '.join(workbook.sheetnames)}"
        )

    # --- Lectura de la tabla de actividades ---

    def _read_activities(
        self,
        sheet: "openpyxl.worksheet.worksheet.Worksheet",
        layout: TableLayout,
    ) -> tuple[tuple[RawActivity, ...], int | None]:
        """Lee la tabla principal; devuelve las actividades y la fila donde paró."""
        activities: list[RawActivity] = []
        stop_row: int | None = None
        current_ordinal: int | None = None
        current_label: str = ""
        current_entregables: list[RawEntregable] = []

        def flush() -> None:
            if current_ordinal is not None:
                activities.append(
                    RawActivity(
                        ordinal=current_ordinal,
                        label=current_label,
                        entregables=tuple(current_entregables),
                    )
                )

        for row in range(layout.header_row + 1, sheet.max_row + 1):
            id_value = sheet.cell(row=row, column=layout.id_col).value
            label_value = sheet.cell(row=row, column=layout.activities_col).value
            content_value = sheet.cell(row=row, column=layout.description_col).value

            # Fin de la tabla de actividades.
            if is_stop_marker(str(id_value or "")) or is_stop_marker(str(label_value or "")):
                stop_row = row
                break

            ordinal = self._parse_ordinal(id_value)

            if ordinal is not None:
                # Comienza una nueva actividad: se cierra la anterior.
                flush()
                current_ordinal = ordinal
                current_label = str(label_value).strip() if label_value is not None else ""
                current_entregables = []

            # Cada sub-fila (inicial o de continuación) es un entregable.
            if current_ordinal is not None:
                entregable_text = self._entregable_text(sheet, row, layout, label_value)
                content_text = str(content_value).strip() if content_value else ""
                if entregable_text or content_text:
                    current_entregables.append(
                        RawEntregable(
                            entregable_text=entregable_text,
                            raw_content=content_text,
                        )
                    )

        flush()
        return tuple(activities), stop_row

    # --- Sección "Otras actividades solicitadas" ---

    def _read_other_activities(
        self,
        sheet: "openpyxl.worksheet.worksheet.Worksheet",
        layout: TableLayout,
        stop_row: int | None,
    ) -> tuple[RawOtherActivity, ...]:
        """Lee las actividades adicionales que siguen a la tabla principal.

        La fila del marcador "OTRAS ACTIVIDADES SOLICITADAS POR..." es a la vez
        el encabezado de la sección: la descripción va en la misma columna del
        marcador y la fecha en la columna "FECHA DE EJECUCIÓN". Si la parada de
        la tabla principal no fue por ese marcador, no hay sección.
        """
        if stop_row is None or not self._is_other_activities_row(sheet, stop_row):
            return ()

        date_col = self._find_other_date_col(sheet, stop_row)
        others: list[RawOtherActivity] = []

        for row in range(stop_row + 1, sheet.max_row + 1):
            id_text = normalize_text(str(sheet.cell(row=row, column=layout.id_col).value or ""))
            text_value = sheet.cell(row=row, column=layout.activities_col).value
            text = str(text_value).strip() if text_value is not None else ""
            if any(marker in id_text or marker in normalize_text(text) for marker in _OTHER_END_MARKERS):
                break
            if not text:
                continue  # filas con ID pero sin descripción: no diligenciadas
            date_value = (
                sheet.cell(row=row, column=date_col).value if date_col is not None else None
            )
            others.append(RawOtherActivity(text=text, date=self._format_date(date_value)))

        return tuple(others)

    @staticmethod
    def _is_other_activities_row(
        sheet: "openpyxl.worksheet.worksheet.Worksheet", row: int
    ) -> bool:
        for col in range(1, sheet.max_column + 1):
            value = normalize_text(str(sheet.cell(row=row, column=col).value or ""))
            if _OTHER_ACTIVITIES_MARKER in value:
                return True
        return False

    @staticmethod
    def _find_other_date_col(
        sheet: "openpyxl.worksheet.worksheet.Worksheet", header_row: int
    ) -> int | None:
        for col in range(1, sheet.max_column + 1):
            value = normalize_text(str(sheet.cell(row=header_row, column=col).value or ""))
            if value.startswith(_OTHER_DATE_HEADER):
                return col
        return None

    @staticmethod
    def _format_date(value: object) -> str:
        if isinstance(value, datetime | date):
            return value.strftime("%d/%m/%Y")
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _entregable_text(
        sheet: "openpyxl.worksheet.worksheet.Worksheet",
        row: int,
        layout: TableLayout,
        label_value: object,
    ) -> str:
        """Texto del entregable de la fila.

        Si la plantilla tiene columna ENTREGABLES, se usa esa. Si no (p. ej. HOCOL),
        la actividad es su propio entregable: se usa el texto de ACTIVIDADES.
        """
        if layout.entregables_col is not None:
            value = sheet.cell(row=row, column=layout.entregables_col).value
            return str(value).strip() if value else ""
        return str(label_value).strip() if label_value else ""

    @staticmethod
    def _parse_ordinal(value: object) -> int | None:
        """Convierte el valor de la columna ID en un entero positivo, o ``None``."""
        if value is None:
            return None
        try:
            number = int(float(str(value).strip()))
        except (ValueError, TypeError):
            return None
        return number if number > 0 else None
